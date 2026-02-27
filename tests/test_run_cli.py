from __future__ import annotations

import json
from pathlib import Path
from uuid import uuid4

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from hoa_report.engine import TEMPLATE_REPORT_COLUMNS
from hoa_report.run import main

_TEST_TMP_DIR = Path("data/_test_tmp")


def _write_config(filename: str, payload: dict[str, object]) -> Path:
    _TEST_TMP_DIR.mkdir(parents=True, exist_ok=True)
    config_path = _TEST_TMP_DIR / filename
    config_path.write_text(json.dumps(payload), encoding="utf-8")
    return config_path


def _write_tape_fixture(filename: str) -> Path:
    return _write_tape_fixture_with_ids(filename, ["L-1001", "L-1002", "L-1003"])


def _write_tape_fixture_with_ids(filename: str, loan_numbers: list[str]) -> Path:
    tape_path = _TEST_TMP_DIR / filename
    df = pd.DataFrame({"Loan Number": loan_numbers})
    df.to_excel(tape_path, index=False)
    return tape_path


def _write_vendor_fixture(filename: str, rows: list[dict[str, object]]) -> Path:
    vendor_path = _TEST_TMP_DIR / filename
    pd.DataFrame(rows).to_excel(vendor_path, index=False)
    return vendor_path


def _write_clayton_vendor_fixture(filename: str, rows: list[dict[str, object]]) -> Path:
    vendor_path = _TEST_TMP_DIR / filename
    with pd.ExcelWriter(vendor_path) as writer:
        pd.DataFrame(rows).to_excel(writer, index=False, sheet_name="HOA")
    return vendor_path


def _write_consolidated_vendor_fixture(filename: str, rows: list[dict[str, object]]) -> Path:
    vendor_path = _TEST_TMP_DIR / filename
    with pd.ExcelWriter(vendor_path) as writer:
        pd.DataFrame(rows).to_excel(writer, index=False, sheet_name="Redwood Additional Data")
    return vendor_path


def _write_template_fixture(filename: str) -> Path:
    template_path = _TEST_TMP_DIR / filename
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    for col_idx, header in enumerate(TEMPLATE_REPORT_COLUMNS, start=1):
        sheet.cell(row=1, column=col_idx, value=header)
    workbook.save(template_path)
    return template_path


def test_cli_runs_end_to_end_with_path_overrides_and_prints_qa_summary(capsys: pytest.CaptureFixture[str]) -> None:
    tape_path = _write_tape_fixture("run_cli.tape.synthetic.xlsx")
    template_path = _write_template_fixture("run_cli.template.synthetic.xlsx")
    vendor_a_path = _write_vendor_fixture(
        "run_cli.vendor_a.synthetic.xlsx",
        [
            {"Loan Number": "L-1001", "Monthly Dues": 125.0},
            {"Loan Number": "L-1002", "Monthly Dues": 225.0},
        ],
    )
    vendor_b_path = _write_vendor_fixture(
        "run_cli.vendor_b.synthetic.xlsx",
        [
            {"Loan Number": "L-1003", "Monthly Dues": 300.0},
        ],
    )

    config_path = _write_config(
        "run_cli.override.config.json",
        {
            "tape_path": "tests/fixtures/does-not-exist.xlsx",
            "template_path": "tests/fixtures/does-not-exist-template.xlsx",
            "vendor_paths": ["tests/fixtures/does-not-exist-vendor.xlsx"],
            "vendor_type": "example_vendor",
            "output_path": "data/_test_tmp/should-not-be-used.xlsx",
        },
    )
    output_path = _TEST_TMP_DIR / f"run_cli.override.output.{uuid4().hex}.xlsx"

    exit_code = main(
        [
            "--config",
            str(config_path),
            "--tape-path",
            str(tape_path),
            "--template-path",
            str(template_path),
            "--vendor-path",
            str(vendor_a_path),
            "--vendor-path",
            str(vendor_b_path),
            "--out",
            str(output_path),
        ]
    )

    captured = capsys.readouterr()
    assert exit_code == 0
    assert output_path.exists()
    assert "Input path validation: OK" in captured.out
    assert "QA Summary" in captured.out
    assert "Match Rate" in captured.out


def test_cli_errors_when_override_path_does_not_exist(capsys: pytest.CaptureFixture[str]) -> None:
    config_path = _write_config(
        "run_cli.invalid_path.config.json",
        {
            "tape_path": "tests/fixtures/tape.synthetic.xlsx",
            "template_path": "tests/fixtures/template.synthetic.xlsx",
            "vendor_paths": ["tests/fixtures/vendor_a.synthetic.xlsx"],
            "vendor_type": "example_vendor",
            "output_path": "data/_test_tmp/run_cli.invalid.output.xlsx",
        },
    )

    with pytest.raises(SystemExit) as exc_info:
        main(
            [
                "--config",
                str(config_path),
                "--vendor-path",
                "tests/fixtures/does-not-exist-vendor.xlsx",
            ]
        )

    captured = capsys.readouterr()
    assert exc_info.value.code == 2
    assert "do not exist" in captured.err


def test_cli_resolves_deal_id_in_output_filename(capsys: pytest.CaptureFixture[str]) -> None:
    tape_path = _write_tape_fixture("run_cli.deal_id.tape.synthetic.xlsx")
    template_path = _write_template_fixture("run_cli.deal_id.template.synthetic.xlsx")
    vendor_path = _write_vendor_fixture(
        "run_cli.deal_id.vendor.synthetic.xlsx",
        [
            {"Loan Number": "L-1001", "Monthly Dues": 125.0},
            {"Loan Number": "L-1002", "Monthly Dues": 225.0},
        ],
    )

    deal_id = f"2026-3-{uuid4().hex[:8]}"
    output_template_path = _TEST_TMP_DIR / "SEMT {deal_id} Servicer HOA.xlsx"
    expected_output_path = _TEST_TMP_DIR / f"SEMT {deal_id} Servicer HOA.xlsx"

    config_path = _write_config(
        "run_cli.deal_id.config.json",
        {
            "tape_path": str(tape_path),
            "template_path": str(template_path),
            "deal_id": deal_id,
            "vendor_paths": [str(vendor_path)],
            "vendor_type": "example_vendor",
            "output_path": str(output_template_path),
        },
    )

    exit_code = main(["--config", str(config_path)])

    captured = capsys.readouterr()
    assert exit_code == 0
    assert expected_output_path.exists()
    assert f"Output written to: {expected_output_path}" in captured.out


def test_cli_runs_sql_enrichment_with_connection_string(
    capsys: pytest.CaptureFixture[str],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    tape_path = _write_tape_fixture("run_cli.sql.tape.synthetic.xlsx")
    template_path = _write_template_fixture("run_cli.sql.template.synthetic.xlsx")
    vendor_a_path = _write_vendor_fixture(
        "run_cli.sql.vendor_a.synthetic.xlsx",
        [
            {"Loan Number": "L-1001", "Monthly Dues": 125.0},
            {"Loan Number": "L-1002", "Monthly Dues": 225.0},
        ],
    )
    output_path = _TEST_TMP_DIR / f"run_cli.sql.output.{uuid4().hex}.xlsx"

    config_path = _write_config(
        "run_cli.sql.config.json",
        {
            "tape_path": str(tape_path),
            "template_path": str(template_path),
            "vendor_paths": [str(vendor_a_path)],
            "vendor_type": "example_vendor",
            "output_path": str(output_path),
            "run_sql": True,
            "sql": {
                "connection_string": (
                    "mssql+pyodbc://@RTSQLGEN01/LOANDATA?"
                    "driver=ODBC+Driver+17+for+SQL+Server&trusted_connection=yes"
                ),
                "query_path": "sql/hoa_enrich.sql",
            },
        },
    )

    calls: list[dict[str, object]] = []

    def _mock_run_sql_enrichment_query(*, tape_df: pd.DataFrame, connection_string: str, query_path: Path) -> pd.DataFrame:
        calls.append(
            {
                "rows": len(tape_df),
                "connection_string": connection_string,
                "query_path": query_path,
            }
        )
        return pd.DataFrame(
            {
                "loan_id": ["L1001", "L1002", "L1003"],
                "Seller": ["SQL Seller A", "SQL Seller B", "SQL Seller C"],
            }
        )

    monkeypatch.setattr("hoa_report.run.run_sql_enrichment_query", _mock_run_sql_enrichment_query)

    exit_code = main(["--config", str(config_path)])

    _ = capsys.readouterr()
    assert exit_code == 0
    assert output_path.exists()
    assert len(calls) == 1
    assert calls[0]["rows"] == 3
    assert calls[0]["query_path"] == Path("sql/hoa_enrich.sql")
    assert str(calls[0]["connection_string"]).startswith("mssql+pyodbc://@RTSQLGEN01/LOANDATA")


def test_cli_errors_with_actionable_message_when_output_is_locked(
    capsys: pytest.CaptureFixture[str],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    tape_path = _write_tape_fixture("run_cli.locked_output.tape.synthetic.xlsx")
    template_path = _write_template_fixture("run_cli.locked_output.template.synthetic.xlsx")
    vendor_path = _write_vendor_fixture(
        "run_cli.locked_output.vendor.synthetic.xlsx",
        [{"Loan Number": "L-1001", "Monthly Dues": 125.0}],
    )
    output_path = _TEST_TMP_DIR / "run_cli.locked_output.xlsx"

    config_path = _write_config(
        "run_cli.locked_output.config.json",
        {
            "tape_path": str(tape_path),
            "template_path": str(template_path),
            "vendor_paths": [str(vendor_path)],
            "vendor_type": "example_vendor",
            "output_path": str(output_path),
        },
    )

    def _raise_permission_error(**kwargs: object) -> Path:
        raise PermissionError(13, "Permission denied", str(kwargs["output_path"]))

    monkeypatch.setattr("hoa_report.run.write_report_from_template", _raise_permission_error)

    with pytest.raises(SystemExit) as exc_info:
        main(["--config", str(config_path)])

    captured = capsys.readouterr()
    assert exc_info.value.code == 2
    assert "permission denied" in captured.err.lower()
    assert "close the file" in captured.err.lower()
    assert "--out" in captured.err


def test_cli_clayton_fills_only_blank_hoa_fields_and_prints_qa_summary(
    capsys: pytest.CaptureFixture[str],
) -> None:
    tape_path = _write_tape_fixture_with_ids(
        "run_cli.clayton.tape.synthetic.xlsx",
        ["L-1001", "L-1002", "L-1003", "L-1004"],
    )
    template_path = _write_template_fixture("run_cli.clayton.template.synthetic.xlsx")
    clayton_path = _write_clayton_vendor_fixture(
        "run_cli.clayton.vendor.synthetic.xlsx",
        [
            {"Loan Number": "L-1001", "HOA Monthly Premium Amount": "$125.00"},
            {"Loan Number": "L-1002", "HOA Monthly Premium Amount": "0"},
            {"Loan Number": "L-1003", "HOA Monthly Premium Amount": None},
            {"Loan Number": "X-9999", "HOA Monthly Premium Amount": "300"},
        ],
    )
    output_path = _TEST_TMP_DIR / f"run_cli.clayton.output.{uuid4().hex}.xlsx"

    config_path = _write_config(
        "run_cli.clayton.config.json",
        {
            "tape_path": str(tape_path),
            "template_path": str(template_path),
            "vendor_paths": [str(clayton_path)],
            "vendor_type": "clayton",
            "output_path": str(output_path),
        },
    )

    exit_code = main(["--config", str(config_path)])

    captured = capsys.readouterr()
    assert exit_code == 0
    assert output_path.exists()
    assert "QA Summary" in captured.out
    assert "Match Rate" in captured.out

    workbook = load_workbook(output_path)
    report_sheet = workbook["Sheet1"]

    hoa_col_idx = TEMPLATE_REPORT_COLUMNS.index("HOA") + 1
    hoa_payment_col_idx = TEMPLATE_REPORT_COLUMNS.index("HOA Monthly Payment") + 1

    assert report_sheet.cell(row=2, column=hoa_col_idx).value == "Y"
    assert report_sheet.cell(row=2, column=hoa_payment_col_idx).value == 125.0

    assert report_sheet.cell(row=3, column=hoa_col_idx).value == "N"
    assert report_sheet.cell(row=3, column=hoa_payment_col_idx).value == 0.0

    assert report_sheet.cell(row=4, column=hoa_col_idx).value in ("", None)
    assert report_sheet.cell(row=4, column=hoa_payment_col_idx).value is None

    assert report_sheet.cell(row=5, column=hoa_col_idx).value in ("", None)
    assert report_sheet.cell(row=5, column=hoa_payment_col_idx).value is None


def test_cli_dd_hoa_defaults_blank_matched_values_to_zero(
    capsys: pytest.CaptureFixture[str],
) -> None:
    tape_path = _write_tape_fixture_with_ids(
        "run_cli.dd_default_zero.tape.synthetic.xlsx",
        ["L-1001", "L-1002", "L-1003"],
    )
    template_path = _write_template_fixture("run_cli.dd_default_zero.template.synthetic.xlsx")
    dd_path = _write_vendor_fixture(
        "run_cli.dd_default_zero.vendor.synthetic.xlsx",
        [
            {"Loan Number": "L-1001", "Monthly HOA Dues ($)": "$125.00"},
            {"Loan Number": "L-1002", "Monthly HOA Dues ($)": None},
            {"Loan Number": "X-9999", "Monthly HOA Dues ($)": "$300.00"},
        ],
    )
    output_path = _TEST_TMP_DIR / f"run_cli.dd_default_zero.output.{uuid4().hex}.xlsx"

    config_path = _write_config(
        "run_cli.dd_default_zero.config.json",
        {
            "tape_path": str(tape_path),
            "template_path": str(template_path),
            "vendor_paths": [str(dd_path)],
            "vendor_type": "dd_hoa",
            "output_path": str(output_path),
        },
    )

    exit_code = main(["--config", str(config_path)])

    captured = capsys.readouterr()
    assert exit_code == 0
    assert output_path.exists()
    assert "QA Summary" in captured.out

    workbook = load_workbook(output_path)
    report_sheet = workbook["Sheet1"]

    hoa_col_idx = TEMPLATE_REPORT_COLUMNS.index("HOA") + 1
    hoa_payment_col_idx = TEMPLATE_REPORT_COLUMNS.index("HOA Monthly Payment") + 1

    assert report_sheet.cell(row=2, column=hoa_col_idx).value == "Y"
    assert report_sheet.cell(row=2, column=hoa_payment_col_idx).value == 125.0

    # Matched DD row has blank HOA dues in source, so output defaults to 0.0 / N.
    assert report_sheet.cell(row=3, column=hoa_col_idx).value == "N"
    assert report_sheet.cell(row=3, column=hoa_payment_col_idx).value == 0.0

    assert report_sheet.cell(row=4, column=hoa_col_idx).value in ("", None)
    assert report_sheet.cell(row=4, column=hoa_payment_col_idx).value is None


def test_cli_limited_review_hardcodes_hoa_fields(
    capsys: pytest.CaptureFixture[str],
) -> None:
    _TEST_TMP_DIR.mkdir(parents=True, exist_ok=True)
    tape_path = _TEST_TMP_DIR / "run_cli.limited_review.tape.synthetic.xlsx"
    pd.DataFrame(
        {
            "Loan Number": ["L-1001", "L-1002", "L-1003"],
            "Review Status": ["Pass", "Limited Review", "Limited Review"],
        }
    ).to_excel(tape_path, index=False)

    template_path = _write_template_fixture("run_cli.limited_review.template.synthetic.xlsx")
    dd_path = _write_vendor_fixture(
        "run_cli.limited_review.vendor.synthetic.xlsx",
        [
            {"Loan Number": "L-1001", "Monthly HOA Dues ($)": "$125.00"},
            {"Loan Number": "L-1002", "Monthly HOA Dues ($)": "$200.00"},
            {"Loan Number": "X-9999", "Monthly HOA Dues ($)": "$300.00"},
        ],
    )
    output_path = _TEST_TMP_DIR / f"run_cli.limited_review.output.{uuid4().hex}.xlsx"

    config_path = _write_config(
        "run_cli.limited_review.config.json",
        {
            "tape_path": str(tape_path),
            "template_path": str(template_path),
            "vendor_paths": [str(dd_path)],
            "vendor_type": "dd_hoa",
            "output_path": str(output_path),
        },
    )

    exit_code = main(["--config", str(config_path)])

    captured = capsys.readouterr()
    assert exit_code == 0
    assert output_path.exists()
    assert "QA Summary" in captured.out

    workbook = load_workbook(output_path)
    report_sheet = workbook["Sheet1"]

    hoa_col_idx = TEMPLATE_REPORT_COLUMNS.index("HOA") + 1
    hoa_payment_col_idx = TEMPLATE_REPORT_COLUMNS.index("HOA Monthly Payment") + 1

    assert report_sheet.cell(row=2, column=hoa_col_idx).value == "Y"
    assert report_sheet.cell(row=2, column=hoa_payment_col_idx).value == 125.0

    assert report_sheet.cell(row=3, column=hoa_col_idx).value == "TBD"
    assert (
        report_sheet.cell(row=3, column=hoa_payment_col_idx).value
        == "Limited Review - please refer to URAR"
    )

    # Override applies even when a Limited Review row has no matched vendor HOA amount.
    assert report_sheet.cell(row=4, column=hoa_col_idx).value == "TBD"
    assert (
        report_sheet.cell(row=4, column=hoa_payment_col_idx).value
        == "Limited Review - please refer to URAR"
    )


def test_cli_consolidated_analytics_defaults_blank_matched_values_to_zero(
    capsys: pytest.CaptureFixture[str],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    tape_path = _write_tape_fixture_with_ids(
        "run_cli.consolidated_default_zero.tape.synthetic.xlsx",
        ["L-1001", "L-1002", "L-1003"],
    )
    template_path = _write_template_fixture("run_cli.consolidated_default_zero.template.synthetic.xlsx")
    consolidated_path = _write_consolidated_vendor_fixture(
        "run_cli.consolidated_default_zero.vendor.synthetic.xlsx",
        [
            {"Loan ID": "C-1", "Monthly HOA Payment Amount": None},
            {"Loan ID": "C-2", "Monthly HOA Payment Amount": "$220.00"},
            {"Loan ID": "C-999", "Monthly HOA Payment Amount": "$50.00"},
        ],
    )
    output_path = _TEST_TMP_DIR / f"run_cli.consolidated_default_zero.output.{uuid4().hex}.xlsx"

    config_path = _write_config(
        "run_cli.consolidated_default_zero.config.json",
        {
            "tape_path": str(tape_path),
            "template_path": str(template_path),
            "vendors": [
                {
                    "name": "consolidated_analytics",
                    "type": "consolidated_analytics",
                    "path": str(consolidated_path),
                    "match_key": "collateral_id",
                }
            ],
            "vendor_priority": ["consolidated_analytics"],
            "output_path": str(output_path),
            "run_sql": True,
            "sql": {
                "connection_string": (
                    "mssql+pyodbc://@RTSQLGEN01/LOANDATA?"
                    "driver=ODBC+Driver+17+for+SQL+Server&trusted_connection=yes"
                ),
                "query_path": "sql/hoa_enrich.sql",
            },
        },
    )

    def _mock_run_sql_enrichment_query(
        *,
        tape_df: pd.DataFrame,
        connection_string: str,
        query_path: Path,
    ) -> pd.DataFrame:
        _ = tape_df, connection_string, query_path
        return pd.DataFrame(
            {
                "loan_id": ["L1001", "L1002", "L1003"],
                "Collateral ID": ["C1", "C2", "C3"],
            }
        )

    monkeypatch.setattr("hoa_report.run.run_sql_enrichment_query", _mock_run_sql_enrichment_query)

    exit_code = main(["--config", str(config_path)])

    captured = capsys.readouterr()
    assert exit_code == 0
    assert output_path.exists()
    assert "QA Summary" in captured.out

    workbook = load_workbook(output_path)
    report_sheet = workbook["Sheet1"]

    hoa_col_idx = TEMPLATE_REPORT_COLUMNS.index("HOA") + 1
    hoa_payment_col_idx = TEMPLATE_REPORT_COLUMNS.index("HOA Monthly Payment") + 1

    # Matched consolidated row has blank HOA dues in source, so output defaults to 0.0 / N.
    assert report_sheet.cell(row=2, column=hoa_col_idx).value == "N"
    assert report_sheet.cell(row=2, column=hoa_payment_col_idx).value == 0.0

    assert report_sheet.cell(row=3, column=hoa_col_idx).value == "Y"
    assert report_sheet.cell(row=3, column=hoa_payment_col_idx).value == 220.0

    assert report_sheet.cell(row=4, column=hoa_col_idx).value in ("", None)
    assert report_sheet.cell(row=4, column=hoa_payment_col_idx).value is None


def test_cli_runs_multi_vendor_with_collateral_id_mapping_and_priority(
    capsys: pytest.CaptureFixture[str],
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    tape_path = _write_tape_fixture_with_ids(
        "run_cli.multi_vendor.tape.synthetic.xlsx",
        ["L-1001", "L-1002", "L-1003"],
    )
    template_path = _write_template_fixture("run_cli.multi_vendor.template.synthetic.xlsx")
    clayton_path = _write_clayton_vendor_fixture(
        "run_cli.multi_vendor.clayton.synthetic.xlsx",
        [
            {"Loan Number": "L-1001", "HOA Monthly Premium Amount": "$125.00"},
            {"Loan Number": "L-1002", "HOA Monthly Premium Amount": "$300.00"},
        ],
    )
    consolidated_path = _write_consolidated_vendor_fixture(
        "run_cli.multi_vendor.consolidated.synthetic.xlsx",
        [
            {"Loan ID": "C-2", "Monthly HOA Payment Amount": "$250.00"},
            {"Loan ID": "C-3", "Monthly HOA Payment Amount": "0"},
            {"Loan ID": "C-999", "Monthly HOA Payment Amount": "50"},
        ],
    )
    output_path = _TEST_TMP_DIR / f"run_cli.multi_vendor.output.{uuid4().hex}.xlsx"

    config_path = _write_config(
        "run_cli.multi_vendor.config.json",
        {
            "tape_path": str(tape_path),
            "template_path": str(template_path),
            "vendors": [
                {
                    "name": "clayton",
                    "type": "clayton",
                    "path": str(clayton_path),
                    "match_key": "loan_id",
                },
                {
                    "name": "consolidated_analytics",
                    "type": "consolidated_analytics",
                    "path": str(consolidated_path),
                    "match_key": "collateral_id",
                },
            ],
            "vendor_priority": ["clayton", "consolidated_analytics"],
            "output_path": str(output_path),
            "run_sql": True,
            "sql": {
                "connection_string": (
                    "mssql+pyodbc://@RTSQLGEN01/LOANDATA?"
                    "driver=ODBC+Driver+17+for+SQL+Server&trusted_connection=yes"
                ),
                "query_path": "sql/hoa_enrich.sql",
            },
        },
    )

    def _mock_run_sql_enrichment_query(
        *,
        tape_df: pd.DataFrame,
        connection_string: str,
        query_path: Path,
    ) -> pd.DataFrame:
        _ = tape_df, connection_string, query_path
        return pd.DataFrame(
            {
                "loan_id": ["L1001", "L1002", "L1003"],
                "Collateral ID": ["C1", "C2", "C3"],
            }
        )

    monkeypatch.setattr("hoa_report.run.run_sql_enrichment_query", _mock_run_sql_enrichment_query)

    exit_code = main(["--config", str(config_path)])

    captured = capsys.readouterr()
    assert exit_code == 0
    assert output_path.exists()
    assert "QA Summary" in captured.out

    workbook = load_workbook(output_path)
    report_sheet = workbook["Sheet1"]

    hoa_col_idx = TEMPLATE_REPORT_COLUMNS.index("HOA") + 1
    hoa_payment_col_idx = TEMPLATE_REPORT_COLUMNS.index("HOA Monthly Payment") + 1

    assert report_sheet.cell(row=2, column=hoa_col_idx).value == "Y"
    assert report_sheet.cell(row=2, column=hoa_payment_col_idx).value == 125.0

    # Clayton has higher priority and should not be overwritten by consolidated.
    assert report_sheet.cell(row=3, column=hoa_col_idx).value == "Y"
    assert report_sheet.cell(row=3, column=hoa_payment_col_idx).value == 300.0

    assert report_sheet.cell(row=4, column=hoa_col_idx).value == "N"
    assert report_sheet.cell(row=4, column=hoa_payment_col_idx).value == 0.0

    assert set(workbook.sheetnames) == {"Sheet1", "QA Summary"}
