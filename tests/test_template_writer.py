from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from hoa_report.engine import TEMPLATE_REPORT_COLUMNS
from hoa_report.io import write_report_from_template

_TEST_TMP_DIR = Path("data/_test_tmp")


def _build_generated_template(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"

    for col_idx, header in enumerate(TEMPLATE_REPORT_COLUMNS, start=1):
        sheet.cell(row=1, column=col_idx, value=header)

    highlighted_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    bordered = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for col_idx in range(1, len(TEMPLATE_REPORT_COLUMNS) + 1):
        cell = sheet.cell(row=2, column=col_idx, value=f"stale-{col_idx}")
        cell.fill = highlighted_fill
        cell.font = Font(name="Calibri", size=11)
        cell.alignment = Alignment(horizontal="left")
        cell.border = bordered

    for row_idx in range(3, 7):
        for col_idx in range(1, len(TEMPLATE_REPORT_COLUMNS) + 1):
            sheet.cell(row=row_idx, column=col_idx, value=f"old-{row_idx}-{col_idx}")

    workbook.save(path)


def _count_non_blank_rows(sheet, start_row: int, col_count: int) -> int:
    count = 0
    for row_idx in range(start_row, sheet.max_row + 1):
        values = [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, col_count + 1)]
        if any(value is not None and str(value).strip() for value in values):
            count += 1
    return count


def test_write_report_from_template_creates_required_sheets_and_preserves_header_and_row_count(
) -> None:
    _TEST_TMP_DIR.mkdir(parents=True, exist_ok=True)
    template_path = _TEST_TMP_DIR / "template.generated.writer.xlsx"
    output_path = _TEST_TMP_DIR / "report.generated.writer.xlsx"
    _build_generated_template(template_path)

    report_rows: list[dict[str, object]] = []
    for idx in range(1, 4):
        row = {column: f"{column}-{idx}" for column in TEMPLATE_REPORT_COLUMNS}
        row["HOA"] = "Y" if idx == 1 else "N"
        row["HOA Monthly Payment"] = float(idx * 100)
        report_rows.append(row)
    report_df = pd.DataFrame(report_rows, columns=TEMPLATE_REPORT_COLUMNS)

    qa_df = pd.DataFrame(
        {
            "metric": ["merged_rows"],
            "value": [len(report_df)],
        }
    )
    exceptions = {
        "vendor_a": {
            "missing_in_vendor": ["L-3", "L-4"],
            "extra_in_vendor": ["X-99"],
        }
    }

    write_report_from_template(
        template_path=template_path,
        output_path=output_path,
        report_df=report_df,
        qa_df=qa_df,
        exceptions=exceptions,
    )

    workbook = load_workbook(output_path)
    assert {
        "Sheet1",
        "QA Summary",
        "Missing in Vendor",
        "Extra in Vendor",
        "Missing in vendor_a",
        "Extra in vendor_a",
    }.issubset(
        set(workbook.sheetnames)
    )

    report_sheet = workbook["Sheet1"]
    headers = [
        report_sheet.cell(row=1, column=col_idx).value
        for col_idx in range(1, len(TEMPLATE_REPORT_COLUMNS) + 1)
    ]
    assert headers == list(TEMPLATE_REPORT_COLUMNS)
    assert _count_non_blank_rows(report_sheet, start_row=2, col_count=len(TEMPLATE_REPORT_COLUMNS)) == len(
        report_df
    )
    assert all(
        report_sheet.cell(row=5, column=col_idx).value is None
        for col_idx in range(1, len(TEMPLATE_REPORT_COLUMNS) + 1)
    )

    style_id_row_2 = report_sheet.cell(row=2, column=1).style_id
    assert style_id_row_2 != 0
    assert report_sheet.cell(row=3, column=1).style_id == style_id_row_2
    assert report_sheet.cell(row=4, column=1).style_id == style_id_row_2
