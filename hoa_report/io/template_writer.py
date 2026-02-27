from __future__ import annotations

from collections.abc import Mapping
from copy import copy
import numbers
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

_HOA_PAYMENT_HEADER = "HOA Monthly Payment"
_HOA_PAYMENT_NUMBER_FORMAT = "0.00"
_VISIBLE_OUTPUT_SHEETS = frozenset({"Sheet1", "QA Summary"})


def _to_excel_value(value: Any) -> Any:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        return value
    return value


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return not value.strip()
    return False


def _row_has_style(sheet: Worksheet, row_idx: int, col_count: int) -> bool:
    return any(sheet.cell(row=row_idx, column=col_idx).style_id != 0 for col_idx in range(1, col_count + 1))


def _find_style_anchor_row(
    sheet: Worksheet,
    start_row: int,
    template_max_row: int,
    col_count: int,
) -> int | None:
    for row_idx in range(start_row, template_max_row + 1):
        if _row_has_style(sheet, row_idx, col_count):
            return row_idx
    return None


def _resolve_style_source_row(
    sheet: Worksheet,
    target_row: int,
    start_row: int,
    template_max_row: int,
    col_count: int,
    style_anchor_row: int | None,
) -> int | None:
    if target_row <= template_max_row and _row_has_style(sheet, target_row, col_count):
        return target_row

    if target_row <= template_max_row:
        for row_idx in range(target_row - 1, start_row - 1, -1):
            if _row_has_style(sheet, row_idx, col_count):
                return row_idx

    return style_anchor_row


def _copy_row_style(
    sheet: Worksheet,
    source_row: int,
    target_row: int,
    col_count: int,
) -> None:
    for col_idx in range(1, col_count + 1):
        source_cell = sheet.cell(row=source_row, column=col_idx)
        target_cell = sheet.cell(row=target_row, column=col_idx)
        target_cell._style = copy(source_cell._style)

    source_dim = sheet.row_dimensions[source_row]
    if source_dim.height is not None:
        sheet.row_dimensions[target_row].height = source_dim.height


def _clear_existing_report_values(
    sheet: Worksheet,
    start_row: int,
    end_row: int,
    col_count: int,
) -> None:
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(1, col_count + 1):
            sheet.cell(row=row_idx, column=col_idx).value = None


def _is_numeric_hoa_payment(value: Any) -> bool:
    if isinstance(value, bool):
        return False
    if not isinstance(value, numbers.Real):
        return False
    return not bool(pd.isna(value))


def _write_dataframe(
    sheet: Worksheet,
    df: pd.DataFrame,
    start_row: int = 1,
    start_col: int = 1,
    default_headers: list[str] | None = None,
) -> None:
    row_idx = start_row
    columns = list(df.columns)
    if not columns and default_headers:
        columns = list(default_headers)

    if columns:
        for col_offset, header in enumerate(columns):
            sheet.cell(row=row_idx, column=start_col + col_offset, value=header)
        row_idx += 1

    if df.empty:
        return

    for row_values in df.itertuples(index=False, name=None):
        for col_offset, value in enumerate(row_values):
            sheet.cell(row=row_idx, column=start_col + col_offset, value=_to_excel_value(value))
        row_idx += 1


def _replace_sheet(workbook: Workbook, title: str) -> Worksheet:
    if title in workbook.sheetnames:
        existing = workbook[title]
        sheet_index = workbook.index(existing)
        workbook.remove(existing)
        return workbook.create_sheet(title=title, index=sheet_index)
    return workbook.create_sheet(title=title)


def _set_sheet_visibility(workbook: Workbook, visible_titles: set[str] | frozenset[str]) -> None:
    for title in workbook.sheetnames:
        sheet = workbook[title]
        sheet.sheet_state = "visible" if title in visible_titles else "hidden"


def _build_exception_rows(
    exceptions: Mapping[str, Mapping[str, Any]],
    key: str,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for vendor, values in exceptions.items():
        raw_loan_ids = values.get(key)
        if isinstance(raw_loan_ids, str) or raw_loan_ids is None:
            continue
        if not isinstance(raw_loan_ids, list | tuple | set):
            continue
        for loan_id in raw_loan_ids:
            if _is_blank(loan_id):
                continue
            rows.append({"Vendor": vendor, "Loan ID": str(loan_id).strip()})
    return pd.DataFrame(rows, columns=["Vendor", "Loan ID"])


def _build_vendor_exception_rows(
    exception_values: Mapping[str, Any],
    key: str,
) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    raw_loan_ids = exception_values.get(key)
    if isinstance(raw_loan_ids, str) or raw_loan_ids is None:
        return pd.DataFrame(rows, columns=["Loan ID"])
    if not isinstance(raw_loan_ids, list | tuple | set):
        return pd.DataFrame(rows, columns=["Loan ID"])

    for loan_id in raw_loan_ids:
        if _is_blank(loan_id):
            continue
        rows.append({"Loan ID": str(loan_id).strip()})
    return pd.DataFrame(rows, columns=["Loan ID"])


def _resolve_vendor_sheet_title(prefix: str, vendor: str, existing_titles: set[str]) -> str:
    base = f"{prefix} {vendor}".strip()
    if len(base) <= 31 and base not in existing_titles:
        existing_titles.add(base)
        return base

    truncated = base[:31]
    if truncated not in existing_titles:
        existing_titles.add(truncated)
        return truncated

    counter = 2
    while True:
        suffix = f" {counter}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        if candidate not in existing_titles:
            existing_titles.add(candidate)
            return candidate
        counter += 1


def write_report_from_template(
    template_path: str | Path,
    output_path: str | Path,
    report_df: pd.DataFrame,
    qa_df: pd.DataFrame | None = None,
    exceptions: Mapping[str, Mapping[str, Any]] | None = None,
) -> Path:
    """Write report output using a template workbook while preserving formatting."""
    template = Path(template_path)
    output = Path(output_path)

    workbook = load_workbook(template)
    if "Sheet1" not in workbook.sheetnames:
        raise ValueError("Template workbook must contain a 'Sheet1' worksheet.")

    report_sheet = workbook["Sheet1"]
    start_row = 2
    report_col_count = int(len(report_df.columns))
    template_max_row = report_sheet.max_row
    hoa_payment_col_offset = (
        int(report_df.columns.get_loc(_HOA_PAYMENT_HEADER))
        if _HOA_PAYMENT_HEADER in report_df.columns
        else None
    )

    if report_col_count:
        _clear_existing_report_values(
            report_sheet,
            start_row=start_row,
            end_row=template_max_row,
            col_count=report_col_count,
        )
        style_anchor_row = _find_style_anchor_row(
            report_sheet,
            start_row=start_row,
            template_max_row=template_max_row,
            col_count=report_col_count,
        )

        for row_offset, row_values in enumerate(report_df.itertuples(index=False, name=None)):
            target_row = start_row + row_offset
            source_row = _resolve_style_source_row(
                report_sheet,
                target_row=target_row,
                start_row=start_row,
                template_max_row=template_max_row,
                col_count=report_col_count,
                style_anchor_row=style_anchor_row,
            )
            if source_row is not None and source_row != target_row:
                _copy_row_style(
                    report_sheet,
                    source_row=source_row,
                    target_row=target_row,
                    col_count=report_col_count,
                )

            for col_offset, value in enumerate(row_values):
                excel_value = _to_excel_value(value)
                target_cell = report_sheet.cell(
                    row=target_row,
                    column=1 + col_offset,
                    value=excel_value,
                )
                if (
                    hoa_payment_col_offset is not None
                    and col_offset == hoa_payment_col_offset
                    and _is_numeric_hoa_payment(excel_value)
                ):
                    target_cell.number_format = _HOA_PAYMENT_NUMBER_FORMAT

    qa_sheet = _replace_sheet(workbook, "QA Summary")
    _write_dataframe(qa_sheet, qa_df if qa_df is not None else pd.DataFrame())

    # Keep exception details in-memory for QA/console reporting only.
    _ = exceptions
    _set_sheet_visibility(workbook, _VISIBLE_OUTPUT_SHEETS)

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    return output
